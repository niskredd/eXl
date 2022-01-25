from openpyxl import workbook, load_workbook
from directory_handling import Diretory_edit

class Automate_excel:

    def __init__(self, data, path, file):
        self.path = path
        self.file = file
        self.data = data

    def save_wb(self):
        self.wb.save(self.path + str(self.p_nr) + " - " + self.data['Addresse'] + ".xlsx")


    def write_to_cell(self, sheet, data, cell):
        ws = self.wb[sheet]
        ws[cell].value = data
        self.wb[sheet] = ws


    def create_sheet(self, name):
        self.wb.create_sheet(name)
        self.dirs_h.content()



    def creat_akonto(self, akonto_data, nr):
        new_akonto = 'Akonto ' + str(nr)

        self.create_sheet(new_akonto)
        ws = self.wb['Kalkyle fakturagrunnlag']
        self.wb[new_akonto] = self.wb.copy_worksheet(ws)

        self.write_to_cell(new_akonto, 1, 'B13')
        self.write_to_cell(new_akonto, beskrivelse, 'C13')
        self.write_to_cell(new_akonto, 1, 'F13')
        self.write_to_cell(new_akonto, 'sum', 'G13')
        self.write_to_cell(new_akonto, sum, 'H13')

        self.save_wb()


    def fillin_sheet(self, prod_dict):
        ws = self.wb['Fylles ut f√∏rst']

        try:

            ws['B4'] = prod_dict['Prosjektnummer']
            ws['B3'] = prod_dict['Skadetype']
            ws['B5'] = prod_dict['Skadenummer']
            ws['B6'] = prod_dict['Selskap']
            ws['B7'] = input(ws['A7'].value + ":  ")
            ws['B8'] = prod_dict['Saksbehandeler']
            ws['B9'] = input(ws['A9'].value + ":  ")
            ws['B10'] = prod_dict['Adresse'] + ' ' + prod_dict['Postnummer'] + ' ' + prod_dict['Poststed']
            ws['B11'] = prod_dict['Adresse'] + ' ' + prod_dict['Postnummer'] + ' ' + prod_dict['Poststed']
            ws['B12'] = prod_dict['Epost']
            ws['B13'] = "Nils Skreddernes"

            ws['B17'] = "1300 Akershus Vest"

            self.save_wb()

        except KeyboardInterrupt:
            self.save_wb()


if __name__ == "__main__":
    project = Automate_excel("1234 - Test")

    project.creat_akonto(str(10000), "Vask", 1)
