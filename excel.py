from openpyxl import workbook, load_workbook
from directory_handling import Diretory_edit

class Automate_excel:

    def __init__(self, state, p_nr, addre):
        if state == 'new':
            self.state = state
            self.wb = load_workbook('Book.xlsx')
            self.p_nr = p_nr
            self.address = addre
            self.dirs_h = Diretory_edit()
            self.dirs_h.change_dir(self.dirs_h.set_dir(str(self.p_nr) + " - " + self.address))
            self.fillin_sheet()
        elif state == 'edit':
            self.dirs_h = Diretory_edit()
            print(self.dirs_h.content())
            
            
        else:
            print("No mach")

    def save_wb(self):
        self.wb.save(str(self.p_nr) + " - " + self.address + ".xlsx")


    def write_to_cell(self, sheet, data, cell):
        ws = self.wb[sheet] 
        ws[cell].value = data
        self.wb[sheet] = ws


    def create_sheet(self, name):
        self.wb.create_sheet(name)
        self.dirs_h.content()
        


    def creat_akonto(self, sum, beskrivelse, nr):
        new_akonto = 'Akonto ' + str(nr)

        self.create_sheet(new_akonto)
        ws = self.wb['Kalkyle fakturagrunnlag']
        self.wb[new_akonto] = self.wb.copy_worksheet(ws)

        self.write_to_cell(new_akonto, 1, 'B13')
        self.write_to_cell(new_akonto, beskrivelse, 'C13')
        self.write_to_cell(new_akonto, 1, 'F13')
        self.write_to_cell(new_akonto, 'sum', 'G13')
        self.write_to_cell(new_akonto, sum, 'H13')

        self.save_wb()


    def fillin_sheet(self):
        ws = self.wb['Fylles ut først']

        try:
            if self.state == 'new':
                ws['B4'] = self.p_nr
            else:
                ws['B4'] = input(ws['A4'].value + ":  ")
            ws['B3'] = input(ws['A3'].value + ":  ")
            ws['B5'] = input(ws['A5'].value + ":  ")
            ws['B6'] = input(ws['A6'].value + ":  ")
            ws['B7'] = input(ws['A7'].value + ":  ")
            ws['B8'] = input(ws['A8'].value + ":  ")
            ws['B9'] = input(ws['A9'].value + ":  ")
            ws['B10'] = input(ws['A10'].value + ":  ")
            ws['B11'] = input(ws['A11'].value + ":  ")
            ws['B12'] = input(ws['A12'].value + ":  ")
            ws['B13'] = "Nils Skreddernes"

            ws['B17'] = "1300 Akershus Vest"

            self.save_wb()

        except KeyboardInterrupt:
            self.save_wb()


if __name__ == "__main__":
    project = Automate_excel("1234 - Test")

    project.creat_akonto(str(10000), "Vask", 1)
